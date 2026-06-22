"""
generate_report.py

분석 결과를 Excel 리포트로 출력한다.
NIQ 공고의 'Reporting & Team Support' 업무(분석 결과 정리 및 리포트 작성)에 대응하는 스크립트.

생성 시트
1. 주간 요약 (Weekly Summary) — 주차별 평균 기분/수면 시간, 이상 감지 건수
2. 이상치 로그 (Anomaly Log) — 이상 감지된 날짜 전체 목록, 조건부 서식으로 강조
3. KPI 추이 (KPI Trend) — 일별 원본 데이터 + 이상 여부

출력 파일: harugyeol_report.xlsx
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import CellIsRule

from db import load_diary
from analyze_anomaly import analyze_anomalies

OUTPUT_PATH = "harugyeol_report.xlsx"

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ANOMALY_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ANOMALY_FONT = Font(color="9C0006")


def _style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4


def build_weekly_summary(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["week"] = df["record_date"].dt.to_period("W").apply(lambda p: p.start_time.date())

    summary = df.groupby("week").agg(
        평균_기분점수=("mood_score", "mean"),
        평균_수면시간=("sleep_duration", "mean"),
        기록_일수=("record_date", "count"),
        이상감지_건수=("anomaly", "sum"),
    ).reset_index()

    summary["평균_기분점수"] = summary["평균_기분점수"].round(2)
    summary["평균_수면시간"] = summary["평균_수면시간"].round(2)
    summary = summary.rename(columns={"week": "주차_시작일"})
    return summary


def write_dataframe(ws, df: pd.DataFrame):
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    _style_header(ws)
    _autofit(ws)


def add_anomaly_highlight(ws, anomaly_col_letter: str, last_row: int):
    rule = CellIsRule(operator="equal", formula=['TRUE'], fill=ANOMALY_FILL, font=ANOMALY_FONT)
    ws.conditional_formatting.add(f"{anomaly_col_letter}2:{anomaly_col_letter}{last_row}", rule)


def generate_report():
    df = load_diary()
    if df.empty:
        print("diary 테이블에 데이터가 없습니다.")
        return

    result = analyze_anomalies(df)

    wb = Workbook()

    # 1. 주간 요약
    ws_weekly = wb.active
    ws_weekly.title = "주간 요약"
    weekly_df = build_weekly_summary(result)
    write_dataframe(ws_weekly, weekly_df)

    # 2. 이상치 로그
    ws_anomaly = wb.create_sheet("이상치 로그")
    anomaly_df = result[result["anomaly"]][
        ["record_date", "anomaly_type", "z_score", "message"]
    ].copy()
    anomaly_df["record_date"] = anomaly_df["record_date"].dt.strftime("%Y-%m-%d")
    if anomaly_df.empty:
        ws_anomaly.append(["이상 감지 없음"])
    else:
        write_dataframe(ws_anomaly, anomaly_df)

    # 3. KPI 추이 (일별 원본 + 이상 여부)
    ws_kpi = wb.create_sheet("KPI 추이")
    kpi_df = result[[
        "record_date", "mood_score", "sleep_duration", "wake_minutes",
        "anomaly", "anomaly_type", "z_score",
    ]].copy()
    kpi_df["record_date"] = kpi_df["record_date"].dt.strftime("%Y-%m-%d")
    write_dataframe(ws_kpi, kpi_df)
    add_anomaly_highlight(ws_kpi, "E", ws_kpi.max_row)

    wb.save(OUTPUT_PATH)
    print(f"리포트 저장 완료: {OUTPUT_PATH}")


if __name__ == "__main__":
    generate_report()
